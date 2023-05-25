import os, sys
import zipfile
import requests
from urllib.parse import urlencode
import argparse
import datetime
# from g import logger


import pandas as pd
import numpy as np
import os, sys, glob
import humanize
import re
import xlrd

import json
import itertools
#from urllib.request import urlopen
#import requests, xmltodict
import time, datetime
import math
from pprint import pprint
import gc
from tqdm import tqdm
tqdm.pandas()
import pickle

import logging
import zipfile
import warnings
import argparse

from sentence_transformers import SentenceTransformer, util

import ipywidgets as widgets
from IPython.display import display
from ipywidgets import Layout, Box, Label

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import units
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment


def load_sentence_model():
    logger.info(f"MultyLangual model dounlowd - start...")
    model = SentenceTransformer('multi-qa-MiniLM-L6-cos-v1')
    # model = SentenceTransformer('all-MiniLM-L6-v2') # хуже - да эмбеддинге-то старые
    logger.info(f"MultyLangual model dounlowd - done!")
    return model


def get_serv_code_name(
        dict_id_score_lst, df_dict_services, 
        dict_code_col_name = 'code', dict_name_col_name = 'name', 
        similarity_threshold=0.9, max_sim_entries=2, max_out_entries=2, 
        debug=False):
    rez= np.array(3*[None])
    if ( (type(dict_id_score_lst)==list) and (len(dict_id_score_lst)>0) and 
        (type(dict_id_score_lst[0])==list) and (len(dict_id_score_lst[0]) > 0) ):
        for i_d, dict_id_score in enumerate(dict_id_score_lst[0]):
            if debug: print("get_code_name: dict_score_id", dict_id_score)
            id, score = dict_id_score.values()
            if debug: print(f"get_code_name: score: {score}, id : {id}")
            if float(score) >= similarity_threshold:
                rez_by_dict_el = df_dict_services.loc[id, [dict_code_col_name, dict_name_col_name]].values
                # print(rez_by_dict_el.shape, rez_by_dict_el)
                if len(rez_by_dict_el.shape) >1:
                    try:
                        rez_by_dict_el = np.insert(rez_by_dict_el, 0, values=round(float(score)*100,1), axis=1)
                    except:
                        rez_by_dict_el = np.insert(rez_by_dict_el, 0, values=round(float(score)*100,1))    
                else:
                    rez_by_dict_el = np.insert(rez_by_dict_el, 0, values=round(float(score)*100,1))
                if i_d == 0:
                    rez = rez_by_dict_el
                else:
                    rez = np.vstack((rez,rez_by_dict_el))
            else: break
        return rez
    else: 
        return np.array([3*[None]])

def semantic_search_serv_by_df (
              df_test_serv, col_name_check, 
              df_dict_services, serv_name_embeddings, 
              model,  
              similarity_threshold=0.9, max_sim_entries=2, max_out_entries=2, 
              n_rows=np.inf,
              debug=False):

    cols_tmplt = ['%', 'code', 'name']
    new_cols_semantic = [ [f"{cols_tmplt[0]}_{il+1:02d}", f"{cols_tmplt[1]}_{il+1:02d}", f"{cols_tmplt[2]}_{il+1:02d}"] for il, lst in enumerate(range(max_out_entries))]
    new_cols_semantic = list(itertools.chain.from_iterable(new_cols_semantic))
    df_test_serv[new_cols_semantic] = None

    max_new_cols = max_out_entries * 3
    empty_values_row = np.array(max_new_cols * [None])
    for i_row, row in tqdm(df_test_serv.iterrows(), total = df_test_serv.shape[0]):
    
        if i_row > n_rows: break

        s = row[col_name_check]
        # id (type(s)==str): # and (len(str) > 0):
        try:
            s_embedding = model.encode(s) 
            dict_id_score_lst = util.semantic_search (s_embedding, serv_name_embeddings, top_k = max_sim_entries)
            values = get_serv_code_name(
                        dict_id_score_lst, df_dict_services,
                        similarity_threshold=similarity_threshold, max_sim_entries=max_sim_entries, max_out_entries=max_out_entries, debug=debug) 
            if len(values)>0:
                if len(values.shape)> 1: # any rows
                    values = list(itertools.chain.from_iterable(values))
                else: values = list(values)
                len_values = len(values)
                if len_values < max_new_cols:
                    values = values + (max_new_cols - len_values) * [None]
                elif len_values > max_new_cols:
                    values = values[:max_new_cols]
                try:
                    df_test_serv.loc[i_row, new_cols_semantic] = values
                    
                except Exception as err:
                    if debug: 
                        print()
                        print("\nsemantic_search:", i_row, s)
                        print("semantic_search:", err, values.shape, values)
                        # print("semantic_search:", err, values.shape, values)
            
            else: 
                df_test_serv.loc[i_row, new_cols_semantic] = empty_values_row
        except Exception as err: 
            print()
            print(err)
            print("semantic_search:", i_row, s)
            df_test_serv.loc[i_row, new_cols_semantic] = empty_values_row
            
    return df_test_serv

def format_excel_sheet_cols(data_processed_dir, fn_xls, format_cols, sheet_name):
    wb = load_workbook(os.path.join(data_processed_dir, fn_xls))
    # ws = wb.active
    ws = wb[sheet_name]
    # l_alignment=Alignment(horizontal='left', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    l_alignment=Alignment(horizontal='left', vertical= 'top', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
    r_alignment=Alignment(horizontal='right', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    border = Border( 
        left=Side(border_style="thin", color='FF000000'),
        right=Side(border_style="thin", color='FF000000'),
        top=Side(border_style="thin", color='FF000000'),
        bottom=Side(border_style="thin", color='FF000000'),
     )
    
    
    # ws.filterMode = True
    last_cell = ws.cell(row=1, column=len(format_cols)) 
    full_range = "A1:" + last_cell.column_letter + str(ws.max_row)
    ws.auto_filter.ref = full_range
    ws.freeze_panes = ws['B2']
    for ic, col_width in enumerate(format_cols):
        cell = ws.cell(row=1, column=ic+1)
        cell.alignment = l_alignment
        ws.column_dimensions[cell.column_letter].width = col_width
    # ft = cell.font
    # ft = Font(bold=False)
    # for row in ws[full_range]: #[1:]
    #     for cell in row:
    #         cell.font = ft    
    #         cell.alignment = l_alignment
    #         cell.border = border
    wb.save(os.path.join(data_processed_dir, fn_xls))

def get_cols_width_exists(ws):
    cols_width_exists = []
    ws.sheet_state, ws.max_row, ws.max_column
    for ic in range(ws.max_column):
        cell = ws.cell(row=1, column=ic+1)
        cols_width_exists.append(ws.column_dimensions[cell.column_letter].width)
    return cols_width_exists
def format_ws_excel_cols(ws, cols_width):
    l_alignment=Alignment(horizontal='left', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    r_alignment=Alignment(horizontal='right', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    # last_cell = ws.cell(row=1, column=len(cols_width)) 
    # full_range = "A1:" + last_cell.column_letter + str(ws.max_row)
    # ws.auto_filter.ref = full_range
    ws.freeze_panes = ws['B2']
    for ic, col_width in enumerate(cols_width):
        cell = ws.cell(row=1, column=ic+1)
        cell.alignment = l_alignment
        ws.column_dimensions[cell.column_letter].width = col_width
    
    return ws  


def rewrite_excel_by_df(
    df_test_serv,
    data_source_dir, data_processed_dir,
    fn_check_file, sheet_name,
    max_sim_entries,
      ):
    wb = load_workbook(os.path.join(data_source_dir, fn_check_file))
    ws = wb[sheet_name]
    cols_width_exists  = get_cols_width_exists(ws)
    # print(cols_width_exists)
    cols_width_new = cols_width_exists
    for _ in range(max_sim_entries):
        cols_width_new += [10., 15., 40.]
    # print(cols_width_new)

    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_save = f"{fn_check_file[:fn_check_file.rfind('.')]}_{str_date}.{fn_check_file.split('.')[-1]}"
    # print(fn_save)
    wb.save(os.path.join(data_processed_dir, fn_save))

    with pd.ExcelWriter(os.path.join(data_processed_dir, fn_save), mode='a', if_sheet_exists='new') as writer: #  engine='openpyxl', 
        # Engine to use for writing. If None, defaults to io.excel.<extension>.writer. NOTE: can only be passed as a keyword argument.
        # Deprecated since version 1.2.0: As the xlwt package is no longer maintained, the xlwt engine will be removed in a future version of pandas.
        # if_sheet_exists{‘error’, ‘new’, ‘replace’, ‘overlay’}, default ‘error’
        df_test_serv.to_excel(writer, sheet_name=f"{sheet_name}_STS", index=False) 
        # работает с Pandas 1.4.1 fail pandas 1.4.4

    format_excel_sheet_cols(data_processed_dir, fn_save, cols_width_new, f"{sheet_name}_STS")

    return fn_save


def semantic_search_serv(
    df_services_804n, serv_name_embeddings,
    fn_check_file_drop_douwn, sheet_name_drop_douwn, col_name_drop_douwn, 
    similarity_threshold_slider, max_entries_slider, max_out_values_slider,
    data_source_dir, data_processed_dir,
    n_rows=np.inf,
    debug=False
    ):
    fn_check_file, sheet_name, col_name_check = fn_check_file_drop_douwn.value, sheet_name_drop_douwn.value, col_name_drop_douwn.value
    similarity_threshold, max_sim_entries, max_out_entries = similarity_threshold_slider.value/100, max_entries_slider.value, max_out_values_slider.value

    if fn_check_file is not None and sheet_name is not None and col_name_check is not None:
        try:
            df_test_serv = pd.read_excel(os.path.join(data_source_dir, fn_check_file), sheet_name = sheet_name)
            display(df_test_serv.head(2))
        except Exception as err:
            df_test_serv = None
            print(str(err))
            print("Укажите в форме выбора парметров корректные значения: файл, лист и колонку для семантического поиска")
            sys.exit(2)
    else:
        print("Укажите в форме выбора парметров: файл, лист и колонку для семантического поиска")
    
    model = load_sentence_model()

    df_test_serv = semantic_search_serv_by_df (
              df_test_serv, col_name_check, 
              df_services_804n, serv_name_embeddings, 
              model,  
              similarity_threshold=similarity_threshold, max_sim_entries=max_sim_entries, max_out_entries=max_out_entries, 
              n_rows=n_rows,
              debug=False)
    display(df_test_serv.head(2))

    logger.info(f"Начато дополнение и форматирование выходного файла...")
    fn_save = rewrite_excel_by_df(    df_test_serv,     data_source_dir, data_processed_dir,
                            fn_check_file, sheet_name,
                            max_sim_entries,
                        )
    logger.info(f"Файл '{fn_save}' сохранен в директорию '{data_processed_dir}'")
    
    return df_test_serv, fn_save

class Logger():
    def __init__(self, name = 'Fuzzy Lookup',
                 strfmt = '[%(asctime)s] [%(levelname)s] > %(message)s', # strfmt = '[%(asctime)s] [%(name)s] [%(levelname)s] > %(message)s'
                 level = logging.INFO,
                 datefmt = '%H:%M:%S', # '%Y-%m-%d %H:%M:%S'
                #  datefmt = '%H:%M:%S %p %Z',

                 ):
        self.name = name
        self.strfmt = strfmt
        self.level = level
        self.datefmt = datefmt
        self.logger = logging.getLogger(name)
        self.logger.setLevel(self.level) #logging.INFO)
        self.offset = datetime.timezone(datetime.timedelta(hours=3))
        # create console handler and set level to debug
        self.ch = logging.StreamHandler()
        self.ch.setLevel(self.level)
        # create formatter
        self.strfmt = strfmt # '[%(asctime)s] [%(levelname)s] > %(message)s'
        self.datefmt = datefmt # '%H:%M:%S'
        # СЃРѕР·РґР°РµРј С„РѕСЂРјР°С‚С‚РµСЂ
        self.formatter = logging.Formatter(fmt=strfmt, datefmt=datefmt)
        self.formatter.converter = lambda *args: datetime.datetime.now(self.offset).timetuple()
        self.ch.setFormatter(self.formatter)
        # add ch to logger
        self.logger.addHandler(self.ch)
logger = Logger().logger
logger.propagate = False

if len(logger.handlers) > 1:
    for handler in logger.handlers:
        logger.removeHandler(handler)
    # del logger
    logger = Logger().logger
    logger.propagate = False

def unzip_file(path_source, fn_zip, work_path):
    logger.info('Unzip ' + fn_zip + ' start...')

    try:
        with zipfile.ZipFile(path_source + fn_zip, 'r') as zip_ref:
            fn_list = zip_ref.namelist()
            zip_ref.extractall(work_path)
        logger.info('Unzip ' + fn_zip + ' done!')
        return fn_list[0]
    except Exception as err:
        logger.error('Unzip error: ' + str(err))
        sys.exit(2)

def save_df_to_excel(df, path_to_save, fn_main, columns = None, b=0, e=None, index=False):
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn = fn_main + '_' + str_date + '.xlsx'
    logger.info(fn + ' save - start ...')
    if e is None or (e <0):
        e = df.shape[0]
    if columns is None:
        df[b:e].to_excel(os.path.join(path_to_save, fn), index = index)
    else:
        df[b:e].to_excel(os.path.join(path_to_save, fn), index = index, columns = columns)
    logger.info(fn + ' saved to ' + path_to_save)
    hfs = get_humanize_filesize(path_to_save, fn)
    logger.info("Size: " + str(hfs))
    return fn

def save_df_lst_to_excel(df_lst, sheet_names_lst, save_path, fn):
    # fn = model + '.xlsx'
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_date = fn.replace('.xlsx','')  + '_' + str_date + '.xlsx'
    
    # with pd.ExcelWriter(os.path.join(path_tkbd_processed, fn_date )) as writer:  
    with pd.ExcelWriter(os.path.join(save_path, fn_date )) as writer:  
        
        for i, df in enumerate(df_lst):
            df.to_excel(writer, sheet_name = sheet_names_lst[i], index=False)
    return fn_date    



def get_humanize_filesize(path, fn):
    human_file_size = None
    try:
        fn_full = os.path.join(path, fn)
    except Exception as err:
        print(err)
        return human_file_size
    if os.path.exists(fn_full):
        file_size = os.path.os.path.getsize(fn_full)
        human_file_size = humanize.naturalsize(file_size)
    return human_file_size
    
def restore_df_from_pickle(path_files, fn_pickle):

    if fn_pickle is None:
        logger.error('Restore pickle from ' + path_files + ' failed!')
        sys.exit(2)
    if os.path.exists(os.path.join(path_files, fn_pickle)):
        df = pd.read_pickle(os.path.join(path_files, fn_pickle))
        # logger.info('Restore ' + re.sub(path_files, '', fn_pickle_СЃ) + ' done!')
        logger.info('Restore ' + fn_pickle + ' done!')
        logger.info('Shape: ' + str(df.shape))
    else:
        # logger.error('Restore ' + re.sub(path_files, '', fn_pickle_СЃ) + ' from ' + path_files + ' failed!')
        logger.error('Restore ' + fn_pickle + ' from ' + path_files + ' failed!')
    return df    



def upload_files_services(supp_dict_dir = '/content/data/supp_dict'):
    base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?'
    # public_key = link #'https://yadi.sk/d/UJ8VMK2Y6bJH7A'  # Сюда вписываете вашу ссылку
    links = [('Коды МГФОМС и 804н.xlsx', 'https://disk.yandex.ru/i/lX1fVnK1J7_hfg', ('МГФОМС', '804н')),
    ('serv_name_embeddings.pk1', 'https://disk.yandex.ru/d/8UTwZg5jKOhxXQ'),
    # ('НВМИ_РМ.xls', 'https://disk.yandex.ru/i/_RotfMJ_cSfeOw', 'Sheet1'),
    # ('МНН.xlsx', 'https://disk.yandex.ru/i/0rMKBimIKbS7ig', 'Sheet1'),
    # ('df_mi_national_release_20230201_2023_02_06_1013.zip', 'https://disk.yandex.ru/d/pfgyT_zmcYrHBw' ),
    # ('df_mi_org_gos_release_20230129_2023_02_07_1331.zip', 'https://disk.yandex.ru/d/Zh-5-FG4uJyLQg' ),
    # ('Специальность (унифицированный).xlsx', 'https://disk.yandex.ru/i/au5M0xyVDW2mtQ', None),
    ]

    # Получаем загрузочную ссылку
    for link_t in links:
        final_url = base_url + urlencode(dict(public_key=link_t[1]))
        response = requests.get(final_url)
        download_url = response.json()['href']

        # Загружаем файл и сохраняем его
        download_response = requests.get(download_url)
        # with open('downloaded_file.txt', 'wb') as f:   # Здесь укажите нужный путь к файлу
        with open(os.path.join(supp_dict_dir, link_t[0]), 'wb') as f:   # Здесь укажите нужный путь к файлу
            f.write(download_response.content)
            logger.info(f"File '{link_t[0]}' uploaded!")
            if link_t[0].split('.')[-1] == 'zip':
                fn_unzip = unzip_file(os.path.join(supp_dict_dir, link_t[0]), '', supp_dict_dir)
                logger.info(f"File '{fn_unzip}' upzipped!")


def load_check_dictionaries_services(path_supp_dicts):
    # global df_services_MGFOMS, df_services_804n, df_RM, df_MNN, df_mi_org_gos, df_mi_national
    # if not os.path.exists(supp_dict_dir):
    #     os.path.mkdir(supp_dict_dir)

    fn = 'Коды МГФОМС.xlsx'
    fn = 'Коды МГФОМС и 804н.xlsx'
    sheet_name = 'МГФОМС'
    df_services_MGFOMS = pd.read_excel(os.path.join(path_supp_dicts, fn), sheet_name = sheet_name)
    df_services_MGFOMS.rename (columns = {'COD': 'code', 'NAME': 'name'}, inplace=True)
    df_services_MGFOMS['code'] = df_services_MGFOMS['code'].astype(str)
    # print("df_services_MGFOMS", df_services_MGFOMS.shape, df_services_MGFOMS.columns)
    logger.info(f"Загружен справочник 'Услуги по реестру  МГФОМС': {str(df_services_MGFOMS.shape)}")

    sheet_name = '804н'
    df_services_804n = pd.read_excel(os.path.join(path_supp_dicts, fn), sheet_name = sheet_name, header=1)
    df_services_804n.rename (columns = {'Код услуги': 'code', 'Наименование медицинской услуги': 'name'}, inplace=True)
    # print("df_services_804n", df_services_804n.shape, df_services_804n.columns)
    logger.info(f"Загружен справочник 'Услуги по приказу 804н': {str(df_services_804n.shape)}")

    fn_pickle = 'serv_name_embeddings.pk1'
    serv_name_embeddings = restore_df_from_pickle(path_supp_dicts, fn_pickle) 
    
    return df_services_MGFOMS, df_services_804n, serv_name_embeddings
# df_services_MGFOMS, df_services_804n, serv_name_embeddings = load_check_dictionaries_services(path_supp_dicts)

def np_unique_nan(lst: np.array, debug = False)->np.array: # a la version 2.4
    lst_unique = None
    if lst is None or (((type(lst)==float) or (type(lst)==np.float64)) and np.isnan(lst)):
        # if debug: print('np_unique_nan:','lst is None or (((type(lst)==float) or (type(lst)==np.float64)) and math.isnan(lst))')
        lst_unique = lst
    else:
        data_types_set = list(set([type(i) for i in lst]))
        if debug: print('np_unique_nan:', 'lst:', lst, 'data_types_set:', data_types_set)
        if ((type(lst)==list) or (type(lst)==np.ndarray)):
            if debug: print('np_unique_nan:','if ((type(lst)==list) or (type(lst)==np.ndarray)):')
            if len(data_types_set) > 1: # несколько типов данных
                if list not in data_types_set and dict not in data_types_set and tuple not in data_types_set and type(None) not in data_types_set:
                    lst_unique = np.array(list(set(lst)), dtype=object)
                else:
                    lst_unique = lst
            elif len(data_types_set) == 1:
                if debug: print("np_unique_nan: elif len(data_types_set) == 1:")
                if list in data_types_set:
                    lst_unique = np.unique(np.array(lst, dtype=object))
                elif  np.ndarray in data_types_set:
                    # print('elif  np.ndarray in data_types_set :')
                    lst_unique = np.unique(lst.astype(object))
                    # lst_unique = np_unique_nan(lst_unique)
                    lst_unique = np.asarray(lst, dtype = object)
                    # lst_unique = np.unique(lst_unique)
                elif type(None) in data_types_set:
                    # lst_unique = np.array(list(set(lst)))
                    lst_unique = np.array(list(set(list(lst))))
                elif dict in  data_types_set:
                    lst_unique = lst
                    # np.unique(lst)
                elif type(lst) == np.ndarray:
                    if debug: print("np_unique_nan: type(lst) == np.ndarray")
                    if (lst.dtype.kind == 'f') or  (lst.dtype == np.float64) or  (float in data_types_set):
                        if debug: print("np_unique_nan: (lst.dtype.kind == 'f')")
                        lst_unique = np.unique(lst.astype(float))
                        # if debug: print("np_unique_nan: lst_unique predfinal:", lst_unique)
                        # lst_unique = np.array(list(set(list(lst))))
                        # if debug: print("np_unique_nan: lst_unique predfinal v2:", lst_unique)
                        # if np.isnan(lst).all():
                        #     lst_unique = np.nan
                        #     if debug: print("np_unique_nan: lst_unique predfinal v3:", lst_unique)
                    elif (lst.dtype.kind == 'S') :
                        if debug: print("np_unique_nan: lst.dtype == string")
                        lst_unique = np.array(list(set(list(lst))))
                        if debug: print(f"np_unique_nan: lst_unique 0: {lst_unique}")
                    elif lst.dtype == object:
                        if debug: print("np_unique_nan: lst.dtype == object")
                        if (type(lst[0])==str) or (type(lst[0])==np.str_) :
                            try:
                                lst_unique = np.unique(lst)
                            except Exception as err:
                                lst_unique = np.array(list(set(list(lst))))
                        else:
                            lst_unique = np.array(list(set(list(lst))))
                        if debug: print(f"np_unique_nan: lst_unique 0: {lst_unique}")
                    else:
                        if debug: print("np_unique_nan: else 0")
                        lst_unique = np.unique(lst)
                else:
                    if debug: print('np_unique_nan:','else i...')
                    lst_unique = np.array(list(set(lst)))
                    
            elif len(data_types_set) == 0:
                lst_unique = None
            else:
                # print('else')
                lst_unique = np.array(list(set(lst)))
        else: # другой тип данных
            if debug: print('np_unique_nan:','другой тип данных')
            # lst_unique = np.unique(np.array(list(set(lst)),dtype=object))
            # lst_unique = np.unique(np.array(list(set(lst)))) # Исходим из того что все елеменыт спсика одного типа
            lst_unique = lst
    if type(lst_unique) == np.ndarray:
        if debug: print('np_unique_nan: final: ', "if type(lst_unique) == np.ndarray")
        if lst_unique.shape[0]==1: 
            if debug: print('np_unique_nan: final: ', "lst_unique.shape[0]==1")
            lst_unique = lst_unique[0]
            if debug: print(f"np_unique_nan: final after: lst_unique: {lst_unique}")
            if (type(lst_unique) == np.ndarray) and (lst_unique.shape[0]==1):  # двойная вложенность
                if debug: print('np_unique_nan: final: ', 'one more', "lst_unique.shape[0]==1")
                lst_unique = lst_unique[0]
        elif lst_unique.shape[0]==0: lst_unique = None
    if debug: print(f"np_unique_nan: return: lst_unique: {lst_unique}")
    if debug: print(f"np_unique_nan: return: type(lst_unique): {type(lst_unique)}")
    return lst_unique


def form_serv_param(fn_list):
    fn_check_file_drop_douwn = widgets.Dropdown( options=fn_list, value=None) #fn_list[0] if len(fn_list) > 0 else None, disabled=False)
    sheet_name_drop_douwn = widgets.Dropdown( options= [None], value= None, disabled=False)
    col_name_drop_douwn = widgets.Dropdown( options= [None], value= None, disabled=False)
    # fn_dict_file_drop_douwn = widgets.Dropdown( options= [None] + fn_list, value= None, disabled=False, )
    # radio_btn_big_dict = widgets.RadioButtons(options=['Да', 'Нет'], value= 'Да', disabled=False) # description='Check me',    , indent=False
    # radio_btn_prod_options = widgets.RadioButtons(options=['Да', 'Нет'], value= 'Нет', disabled=False if radio_btn_big_dict.value=='Да' else True )
    similarity_threshold_slider = widgets.IntSlider(min=1,max=100, value=90)
    max_entries_slider = widgets.IntSlider(min=1,max=5, value=4)
    max_out_values_slider = widgets.IntSlider(min=1,max=10, value=4)

    form_item_layout = Layout(display='flex', flex_flow='row', justify_content='space-between')
    check_box = Box([Label(value='Проверяемый файл:'), fn_check_file_drop_douwn], layout=form_item_layout) 
    sheet_box = Box([Label(value='Имя листа:'), sheet_name_drop_douwn], layout=form_item_layout) 
    column_box = Box([Label(value='Заголовок колонки:'), col_name_drop_douwn], layout=form_item_layout) 
    # big_dict_box = Box([Label(value='Использовать большие справочники:'), radio_btn_big_dict], layout=form_item_layout) 
    # prod_options_box = Box([Label(value='Искать в Вариантах исполнения (+10 мин):'), radio_btn_prod_options], layout=form_item_layout) 
    similarity_threshold_box = Box([Label(value='Минимальный % сходства позиций:'), similarity_threshold_slider], layout=form_item_layout) 
    max_entries_box = Box([Label(value='Максимальное кол-во найденных позиций:'), max_entries_slider], layout=form_item_layout) 
    max_out_values_box = Box([Label(value='Максимальное кол-во выводимых позиций:'), max_out_values_slider], layout=form_item_layout) 
    
    # form_items = [check_box, dict_box, big_dict_box, prod_options_box, similarity_threshold_box, max_entries_box]
    form_items = [check_box, sheet_box, column_box, similarity_threshold_box, max_entries_box, max_out_values_box]
    
    form = Box(form_items, layout=Layout(display='flex', flex_flow= 'column', border='solid 2px', align_items='stretch', width='50%')) #width='auto'))
    # return form, fn_check_file_drop_douwn, fn_dict_file_drop_douwn, radio_btn_big_dict, radio_btn_prod_options, similarity_threshold_slider, max_entries_slider
    return form, fn_check_file_drop_douwn, sheet_name_drop_douwn, col_name_drop_douwn, similarity_threshold_slider, max_entries_slider, max_out_values_slider
    # form = Box(form_items, layout=Layout(display='flex', flex_flow= 'column', border='solid 2px', align_items='stretch', width='70%')) #width='auto'))

def on_fn_check_file_drop_douwn_change(change):
    global sheet_name_drop_douwn, data_source_dir
    # print(change.new)
    xl = pd.ExcelFile(os.path.join(data_source_dir, change.new))
    # xl = pd.ExcelFile(os.path.join('/content/data/source', change.new))
    sheet_lst = list(xl.sheet_names)
    # print(sheet_lst)
    sheet_name_drop_douwn.options = sheet_lst
    sheet_lst_serv = [sheet for sheet in sheet_lst if 'услуг' in sheet.lower()]
    sheet_name_drop_douwn.value = sheet_lst_serv[0] if len(sheet_lst_serv)>0 else sheet_lst[0]
def on_sheet_name_drop_douwn_change(change):
    global col_name_drop_douwn, data_source_dir, fn_check_file_drop_douwn
    df = pd.read_excel(os.path.join(data_source_dir, fn_check_file_drop_douwn.value), sheet_name = change.new)
    # df = pd.read_excel(os.path.join('/content/data/source', fn_check_file_drop_douwn.value), sheet_name = change.new)
    columns_lst = list(df.columns)
    col_name_drop_douwn.options = columns_lst
    columns_lst_serv_name = [col for col in columns_lst if ('наименован' in col.lower() and 'услуг' in col.lower())]
    columns_lst_serv = [col for col in columns_lst if 'услуг' in col.lower()]
    col_name_drop_douwn.value = columns_lst_serv_name[0] if len(columns_lst_serv_name)>0 else (columns_lst_serv[0] if len(columns_lst_serv)>0 else columns_lst[0])
